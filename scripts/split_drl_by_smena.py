"""
DRL Defect Details Excel → Smenalarga ajratish
Foydalanish: py -3 scripts/split_drl_by_smena.py "C:/Users/user/Desktop/drl aprel.xlsx"

Aprel 2026 smena jadvali (tasdiqlangan):
  (kun, E/N) → smena (A / B / D)
"""
import sys
import os
import openpyxl
from collections import defaultdict

# ─── Aprel 2026 smena jadvali ────────────────────────────────────────────────
# (kun, smena_tip) → smena_label
SCHEDULE: dict[tuple[int, str], str] = {
    (1,  'E'): 'D',  (1,  'N'): 'D',
    (3,  'E'): 'D',  (3,  'N'): 'A',
    (4,  'E'): 'D',  (4,  'N'): 'A',
    (5,  'E'): 'B',  (5,  'N'): 'D',
    (6,  'E'): 'B',  (6,  'N'): 'D',
    (7,  'E'): 'B',  (7,  'N'): 'D',
    (8,  'E'): 'A',  (8,  'N'): 'B',
    (9,  'E'): 'A',  (9,  'N'): 'B',
    (10, 'E'): 'A',  (10, 'N'): 'B',
    (11, 'E'): 'D',  (11, 'N'): 'A',
    (12, 'E'): 'A',  (12, 'N'): 'D',   # A=kunduzgi, D=tungi
    (13, 'E'): 'D',  (13, 'N'): 'A',
    (14, 'E'): 'B',  (14, 'N'): 'D',
    (15, 'E'): 'B',  (15, 'N'): 'D',
    (16, 'E'): 'B',  (16, 'N'): 'D',
    (17, 'E'): 'A',  (17, 'N'): 'B',
    (18, 'E'): 'A',  (18, 'N'): 'B',
    (19, 'N'): 'B',  # 19-aprel dam olish; N satrlar (00:xx) = 18-aprel kechasiga tegishli = B smena
    (20, 'E'): 'A',  (20, 'N'): 'D',
    (21, 'E'): 'B',  (21, 'N'): 'D',
    (22, 'E'): 'D',  (22, 'N'): 'A',   # D=kunduzgi, A=tungi
    (23, 'E'): 'D',  (23, 'N'): 'A',
    (24, 'E'): 'D',  (24, 'N'): 'A',
    (25, 'E'): 'B',  (25, 'N'): 'D',   # B=kunduzgi, D=tungi
    (26, 'E'): 'B',  (26, 'N'): 'D',
    (27, 'E'): 'A',  (27, 'N'): 'D',
    (28, 'E'): 'A',  (28, 'N'): 'B',
    (29, 'E'): 'A',  (29, 'N'): 'B',
    (30, 'E'): 'B',  (30, 'N'): 'D',
}


def get_smena(row_values: list) -> str:
    """Defect Date (col 13) va Смена (col 11) → smena label (A/B/D)"""
    defect_date = row_values[13]
    smena_col   = str(row_values[11] or '').strip().upper()
    if not defect_date or not smena_col:
        return 'UNKNOWN'
    day = defect_date.day if hasattr(defect_date, 'day') else None
    if day is None:
        return 'UNKNOWN'
    return SCHEDULE.get((day, smena_col), 'UNKNOWN')


def split_excel(src_path: str):
    print(f'Oqilmoqda: {src_path}')
    wb_src = openpyxl.load_workbook(src_path)
    ws_src = wb_src.active

    all_rows = list(ws_src.iter_rows(values_only=False))
    # Rows 1-5 (index 0-4) → header rows, data starts from index 5
    header_rows = all_rows[:5]
    data_rows   = all_rows[5:]

    # Sort data rows by smena
    buckets: dict[str, list] = defaultdict(list)
    unknown = 0

    for row in data_rows:
        vals = [cell.value for cell in row]
        if not vals or vals[0] != 'S':
            continue  # Skip non-S and empty rows
        label = get_smena(vals)
        if label == 'UNKNOWN':
            unknown += 1
            continue
        buckets[label].append(row)

    print(f'  A smena: {len(buckets["A"])} qator')
    print(f'  B smena: {len(buckets["B"])} qator')
    print(f'  D smena: {len(buckets["D"])} qator')
    print(f'  Unknown: {unknown} qator')

    src_dir  = os.path.dirname(src_path)
    src_base = os.path.splitext(os.path.basename(src_path))[0]

    for label, rows in buckets.items():
        out_path = os.path.join(src_dir, f'{src_base}_{label}.xlsx')

        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.title = ws_src.title

        # Copy column dimensions
        for col_letter, dim in ws_src.column_dimensions.items():
            ws_out.column_dimensions[col_letter].width = dim.width

        # Copy header rows (rows 1-5)
        for hrow in header_rows:
            ws_out.append([cell.value for cell in hrow])

        # Copy data rows for this smena
        for drow in rows:
            ws_out.append([cell.value for cell in drow])

        wb_out.save(out_path)
        print(f'  Yozildi: {out_path}  ({len(rows)} qator)')

    print('Tayyor!')


if __name__ == '__main__':
    path = sys.argv[1] if len(sys.argv) > 1 else r'C:\Users\user\Desktop\drl aprel.xlsx'
    split_excel(path)
